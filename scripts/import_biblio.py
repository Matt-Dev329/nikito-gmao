#!/usr/bin/env python3
"""
Import de la bibliothèque points de contrôle Excel → SQL pour Supabase.

Lit le fichier 02_bibliotheque_points_controle.xlsx (onglet "Bibliothèque")
et génère un fichier SQL d'INSERT prêt à exécuter dans le SQL Editor Supabase.

Le script résout les categorie_id à partir du nom de catégorie via une
sous-requête (SELECT id FROM categories_equipement WHERE nom = ...).

Usage:
    python3 scripts/import_biblio.py 02_bibliotheque_points_controle.xlsx > scripts/import_biblio.sql

Puis dans Supabase SQL Editor:
    Copier-coller le contenu de scripts/import_biblio.sql et exécuter.

Vérification post-import:
    SELECT type_controle, COUNT(*) FROM bibliotheque_points GROUP BY type_controle;
"""
import sys
import openpyxl
from pathlib import Path


def escape_sql(value: str) -> str:
    """Échappe les apostrophes pour SQL."""
    return value.replace("'", "''")


def normalize_bool(v) -> str:
    """Convertit 'OUI'/'non'/None en TRUE/FALSE SQL."""
    if v is None:
        return "FALSE"
    s = str(v).strip().lower()
    if s in ("oui", "yes", "true", "1"):
        return "TRUE"
    return "FALSE"


def normalize_text(v) -> str:
    """Convertit None ou '—' en NULL SQL, sinon string échappée."""
    if v is None:
        return "NULL"
    s = str(v).strip()
    if s in ("", "—", "-"):
        return "NULL"
    return f"'{escape_sql(s)}'"


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 import_biblio.py <fichier.xlsx>", file=sys.stderr)
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"Fichier introuvable: {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Bibliothèque" not in wb.sheetnames:
        print(f"Onglet 'Bibliothèque' absent. Onglets disponibles: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb["Bibliothèque"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if r[0] is not None]

    print("-- " + "=" * 70)
    print("-- Import bibliothèque points de contrôle")
    print(f"-- Source: {path.name}")
    print(f"-- {len(rows)} points à importer")
    print("-- " + "=" * 70)
    print()
    print("BEGIN;")
    print()

    # Statistiques
    types_count = {}
    cats_count = {}
    for r in rows:
        t = r[3] or "?"
        c = r[1] or "?"
        types_count[t] = types_count.get(t, 0) + 1
        cats_count[c] = cats_count.get(c, 0) + 1

    print("-- Stats par type de contrôle:")
    for t, n in sorted(types_count.items()):
        print(f"--   {t}: {n}")
    print(f"-- Stats par catégorie ({len(cats_count)} catégories):")
    for c, n in sorted(cats_count.items()):
        print(f"--   {c}: {n}")
    print()

    for r in rows:
        (
            id_excel,
            categorie,
            libelle,
            type_ctrl,
            assigne_a,
            obligation_const,
            norme,
            bloquant_ko,
            photo_obl,
            ordre,
            verrouille,
        ) = r

        cat_lookup = f"(SELECT id FROM categories_equipement WHERE nom = '{escape_sql(str(categorie))}')"

        cols = [
            "categorie_id",
            "libelle",
            "type_controle",
            "assigne_a",
            "obligation_constructeur",
            "norme_associee",
            "bloquant_si_ko",
            "photo_obligatoire",
            "ordre",
            "verrouille",
        ]
        vals = [
            cat_lookup,
            f"'{escape_sql(str(libelle))}'",
            f"'{type_ctrl}'",
            normalize_text(assigne_a),
            normalize_bool(obligation_const),
            normalize_text(norme),
            normalize_bool(bloquant_ko),
            normalize_bool(photo_obl),
            str(ordre) if ordre is not None else "NULL",
            normalize_bool(verrouille),
        ]
        print(
            f"INSERT INTO bibliotheque_points ({', '.join(cols)}) "
            f"VALUES ({', '.join(vals)});"
        )

    print()
    print("COMMIT;")
    print()
    print("-- Vérification post-import:")
    print("-- SELECT type_controle, COUNT(*) FROM bibliotheque_points GROUP BY type_controle;")


if __name__ == "__main__":
    main()
